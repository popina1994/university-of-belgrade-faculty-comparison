from openpyxl.worksheet import worksheet


class Work:
    def __init__(self, title: str, year: int, authors: str, doc_type: str, author: str = "", num_citations: int=0,
                 document_name: str="", impact_factor: str="", impact_factor5: str=""):
        self._author = author
        self._title = title
        self._year = year
        self._authors = authors
        self._doc_type = doc_type
        self._num_citations = num_citations
        self._document_name = document_name
        self._impact_factor = impact_factor
        self._impact_factor5 = impact_factor5

    COLUMN_AUTHOR = "Glavni autor"
    COLUMN_TITLE = "Naslov"
    COLUMN_YEAR= "Godina"
    COLUMN_AUTHORS= "Autori"
    COLUMN_NUM_CITATIONS = "Broj citiranja"
    COLUMN_DOC_TYPE = "Tip rada"
    COLUMN_DOC_NAME = "Ime dokumenta"
    COLUMN_IMPACT_FACTOR = "Impact factor 2017"
    COLUMN_IMPACT_FACTOR5 = "Impact factor 5 year 2017"

    COLUMN_IDX_AUTHOR = 1
    COLUMN_IDX_TITLE = 2
    COLUMN_IDX_YEAR = 3
    COLUMN_IDX_AUTHORS = 4
    COLUMN_IDX_NUM_CITATIONS = 5
    COLUMN_IDX_DOC_TYPE = 6
    COLUMN_IDX_DOC_NAME = 7
    COLUMN_IDX_IMPACT_FACTOR = 8
    COLUMN_IDX_IMPACT_5FACTOR = 9

    @property
    def title(self):
        return self._title

    @title.setter
    def title(self, value):
        self._title = value

    @property
    def authors(self):
        return self._authors

    @authors.setter
    def authors(self, value):
        self._authors = value

    @property
    def author(self):
        return self._author

    @author.setter
    def author(self, value):
        self._author = value

    @property
    def year(self):
        return self._year

    @year.setter
    def year(self, value):
        self._year = value

    @property
    def doc_type(self):
        return self._doc_type

    @doc_type.setter
    def doc_type(self, value):
        self._doc_type = value

    @property
    def num_citations(self):
        return self._num_citations

    @num_citations.setter
    def num_citations(self, value):
        self._num_citations = value

    @property
    def document_name(self):
        return self._document_name

    @document_name.setter
    def document_name(self, value):
        self._document_name = value

    @property
    def impact_factor(self):
        return self._impact_factor

    @impact_factor.setter
    def impact_factor(self, value):
        self._impact_factor = value

    @property
    def impact_factor5(self):
        return self._impact_factor5

    @impact_factor5.setter
    def impact_factor5(self, value):
        self._impact_factor5 = value

    @staticmethod
    def write_headers_to_sheet(sheet: worksheet):
        sheet.cell(1, Work.COLUMN_IDX_AUTHOR).value = Work.COLUMN_AUTHOR
        sheet.cell(1, Work.COLUMN_IDX_TITLE).value = Work.COLUMN_TITLE
        sheet.cell(1, Work.COLUMN_IDX_YEAR).value = Work.COLUMN_YEAR
        sheet.cell(1, Work.COLUMN_IDX_AUTHORS).value = Work.COLUMN_AUTHORS
        sheet.cell(1, Work.COLUMN_IDX_NUM_CITATIONS).value = Work.COLUMN_NUM_CITATIONS
        sheet.cell(1, Work.COLUMN_IDX_DOC_TYPE).value = Work.COLUMN_DOC_TYPE
        sheet.cell(1, Work.COLUMN_IDX_DOC_NAME).value = Work.COLUMN_DOC_NAME
        sheet.cell(1, Work.COLUMN_IDX_IMPACT_FACTOR).value = Work.COLUMN_IMPACT_FACTOR
        sheet.cell(1, Work.COLUMN_IDX_IMPACT_5FACTOR).value = Work.COLUMN_IMPACT_FACTOR5

    def write_to_sheet(self, sheet: worksheet, row: int):
        sheet.cell(row, Work.COLUMN_IDX_AUTHOR).value = self.author
        sheet.cell(row, Work.COLUMN_IDX_TITLE).value = self.title
        sheet.cell(row, Work.COLUMN_IDX_YEAR).value = self.year
        sheet.cell(row, Work.COLUMN_IDX_AUTHORS).value = self.authors
        sheet.cell(row, Work.COLUMN_IDX_NUM_CITATIONS).value = self.num_citations
        sheet.cell(row, Work.COLUMN_IDX_DOC_TYPE).value = self.doc_type
        sheet.cell(row, Work.COLUMN_IDX_DOC_NAME).value = self.document_name
        sheet.cell(row, Work.COLUMN_IDX_IMPACT_FACTOR).value = self.impact_factor
        sheet.cell(row, Work.COLUMN_IDX_IMPACT_5FACTOR).value = self.impact_factor5
