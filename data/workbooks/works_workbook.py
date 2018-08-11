from enum import IntEnum

from openpyxl import Workbook, load_workbook
from data.tables.work_table.work import Work
from data.tables.work_table.work_scopus import WorkScopus
from data.tables.work_table.work_wos import WorkWos
from utilities.global_setup import DATA_PATH

WORKS_WOS_FILE_NAME = DATA_PATH + r"\Work\works_wos.xlsx"
WORKS_SCOPUS_FILE_NAME = DATA_PATH + r"\Work\works_scopus.xlsx"
WORKS_TEMP_FILE_NAME = DATA_PATH + r"\Work\works_temp.xlsx"

WORKS_FILE_NAME = [WORKS_WOS_FILE_NAME, WORKS_SCOPUS_FILE_NAME, WORKS_TEMP_FILE_NAME]
WORKS_TEMP_WORK_BOOK = WORKS_FILE_NAME.index(WORKS_TEMP_FILE_NAME)
WORKS_SHEET_NAME = "Radovi"


class WorkTypes(IntEnum):
    WOS = 0
    SCOPUS = 1
    TEMPORARY = 2


class WorksWorkbook:
    def __init__(self, work_book_type: WorkTypes, is_write):
        self.work_book_name = WORKS_FILE_NAME[work_book_type]
        self.work_book_type = work_book_type
        self.work_book = Workbook() if is_write else load_workbook(filename=self.work_book_name)
        if is_write:
            self.work_book.remove(self.work_book.active)
            self.row = 2
        self.sheet = self.work_book.create_sheet(WORKS_SHEET_NAME) if is_write else self.work_book[WORKS_SHEET_NAME]
        self.sheet_loaded = False
        self.rows = []

    def save_work(self, work: Work):
        work.write_to_sheet(self.sheet, self.row)
        if self.row == 2:
            work.write_headers_to_sheet(self.sheet)
        self.row += 1

    def load_sheet(self):
        function_read = WorkWos.read_from_sheet if self.work_book_type == WorkTypes.WOS else WorkScopus.read_from_sheet
        for row_idx in range(2, self.sheet.max_row+1):
            self.rows.append(function_read(self.sheet, row_idx))
        return self.rows

    def read_work(self, row: int):
        if not self.sheet_loaded:
            self.load_sheet()
            self.sheet_loaded = True
        return self.rows[row]

    def save(self):
        self.work_book.save(self.work_book_name)
