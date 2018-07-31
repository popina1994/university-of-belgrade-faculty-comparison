from openpyxl import Workbook
from data.author import Author
from utilities.global_setup import DATA_PATH
AUTHORS_FILE_NAME = DATA_PATH + r"\People\authors.xlsx"


class AuthorsWorkbook:
    def __init__(self):
        self.work_book = Workbook()
        self.work_book.remove_sheet(self.work_book.active)

    def save_authors(self, authors, faculty_name):
        work_sheet = self.work_book.create_sheet(faculty_name)
        work_sheet.title = faculty_name

        work_sheet.cell(1, Author.COLUMN_IDX_FIRST_NAME).value = Author.COLUMN_FIRST_NAME
        work_sheet.cell(1, Author.COLUMN_IDX_MIDDLE_NAME).value = Author.COLUMN_MIDDLE_NAME
        work_sheet.cell(1, Author.COLUMN_IDX_LAST_NAME).value = Author.COLUMN_LAST_NAME
        work_sheet.cell(1, Author.COLUMN_IDX_DEPARTMENT_NAME).value = Author.COLUMN_DEPARTMENT_NAME
        work_sheet.cell(1, Author.COLUMN_IDX_FACULTY_NAME).value = Author.COLUMN_FACULTY_NAME
        for row, author in enumerate(authors, start=2):
            work_sheet.cell(row, Author.COLUMN_IDX_FIRST_NAME).value = author.first_name
            work_sheet.cell(row, Author.COLUMN_IDX_MIDDLE_NAME).value = author.middle_name
            work_sheet.cell(row, Author.COLUMN_IDX_LAST_NAME).value = author.last_name
            work_sheet.cell(row, Author.COLUMN_IDX_DEPARTMENT_NAME).value = author.department
            work_sheet.cell(row, Author.COLUMN_IDX_FACULTY_NAME).value = author.faculty

    def save(self):
        self.work_book.save(AUTHORS_FILE_NAME)
