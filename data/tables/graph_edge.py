from openpyxl.worksheet import worksheet


class GraphEdge:
    def __init__(self, author1: str, author2: str, weight: float):
        self._author1 = author1
        self._author2 = author2
        self._weight = weight

    @property
    def author1(self):
        return self._author1

    @author1.setter
    def author1(self, value):
        self._author1 = value

    @property
    def author2(self):
        return self._author2

    @author2.setter
    def author2(self, value):
        self._author2 = value

    @property
    def weight(self):
        return self._weight

    @weight.setter
    def weight(self, value):
        self._weight = value

    COLUMN_IDX_AUTHOR1_NAME = 1
    COLUMN_IDX_AUTHOR2_NAME = 2
    COLUMN_IDX_WEIGHT_NAME = 3
    COLUMN_AUTHOR1_NAME = "Source"
    COLUMN_AUTHOR2_NAME = "Target"
    COLUMN_WEIGHT_NAME = "Weight"

    @staticmethod
    def write_headers_to_sheet(sheet: worksheet):
        sheet.cell(1, GraphEdge.COLUMN_IDX_AUTHOR1_NAME).value = GraphEdge.COLUMN_AUTHOR1_NAME
        sheet.cell(1, GraphEdge.COLUMN_IDX_AUTHOR2_NAME).value = GraphEdge.COLUMN_AUTHOR2_NAME
        sheet.cell(1, GraphEdge.COLUMN_IDX_WEIGHT_NAME).value = GraphEdge.COLUMN_WEIGHT_NAME

    def write_to_sheet(self, sheet: worksheet, row: int):
        sheet.cell(row, GraphEdge.COLUMN_IDX_AUTHOR1_NAME).value = self.author1
        sheet.cell(row, GraphEdge.COLUMN_IDX_AUTHOR2_NAME).value = self.author2
        sheet.cell(row, GraphEdge.COLUMN_IDX_WEIGHT_NAME).value = self.weight
