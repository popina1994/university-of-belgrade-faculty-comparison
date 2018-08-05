from openpyxl.worksheet import worksheet

from data.tables.work_table.work import Work


class WorkScopus(Work):
    COLUMN_WEIGHT_INDEX= "Field-Weighted Citation Impact"
    COLUMN_CITE_SCORE= "CiteScore 2017"
    COLUMN_SJR = "SCImago Journal Rank 2017"
    COLUMN_SNIP = "Source Normalized Impact per Paper 2017"

    COLUMN_IDX_WEIGHT_INDEX = 7
    COLUMN_IDX_CITE_SCORE = 8
    COLUMN_IDX_SJR = 11
    COLUMN_IDX_SNIP = 12

    def __init__(self, title: str, year: int, authors: str, doc_type: str, author: str = "", num_citations: int=0,
                 document_name: str="", department: str="", faculty: str="", weight_index="", cite_score="",
                 sjr="", snip=""):

        super().__init__(title, year, authors, doc_type, author, num_citations, document_name, department, faculty)
        self._cite_score = cite_score
        self._weight_index = weight_index
        self._sjr = sjr
        self._snip = snip

    @property
    def cite_score(self):
        return self._cite_score

    @cite_score.setter
    def cite_score(self, value):
        self._cite_score = value

    @property
    def weight_index(self):
        return self._weight_index

    @weight_index.setter
    def weight_index(self, value):
        self._weight_index = value

    @property
    def sjr(self):
        return self._sjr

    @sjr.setter
    def sjr(self, value):
        self._sjr = value

    @property
    def snip(self):
        return self._snip

    @snip.setter
    def snip(self, value):
        self._snip = value

    @staticmethod
    def write_headers_to_sheet(sheet: worksheet):
        Work.write_headers_to_sheet(sheet)
        sheet.cell(1, WorkScopus.COLUMN_IDX_CITE_SCORE).value = WorkScopus.COLUMN_CITE_SCORE
        sheet.cell(1, WorkScopus.COLUMN_IDX_WEIGHT_INDEX).value = WorkScopus.COLUMN_WEIGHT_INDEX
        sheet.cell(1, WorkScopus.COLUMN_IDX_SJR).value = WorkScopus.COLUMN_SJR
        sheet.cell(1, WorkScopus.COLUMN_IDX_SNIP).value = WorkScopus.COLUMN_SNIP

    def write_to_sheet(self, sheet: worksheet, row: int):
        super().write_to_sheet(sheet, row)
        sheet.cell(row, WorkScopus.COLUMN_IDX_CITE_SCORE).value = self.cite_score
        sheet.cell(row, WorkScopus.COLUMN_IDX_WEIGHT_INDEX).value = self.weight_index
        sheet.cell(row, WorkScopus.COLUMN_IDX_SJR).value = self.sjr
        sheet.cell(row, WorkScopus.COLUMN_IDX_SNIP).value = self.snip
