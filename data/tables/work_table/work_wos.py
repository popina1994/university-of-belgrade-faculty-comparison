from openpyxl.worksheet import worksheet

from data.tables.work_table.work import Work


class WorkWos(Work):
    COLUMN_IMPACT_FACTOR = "Impact factor 2017"
    COLUMN_IMPACT_FACTOR5 = "Impact factor 5 godina 2017"

    COLUMN_IDX_IMPACT_FACTOR = 7
    COLUMN_IDX_IMPACT_FACTOR5 = 8

    def __init__(self, title: str, year: int, authors: str, doc_type: str, author: str = "", num_citations: int=0,
                 document_name: str="", impact_factor: str="", impact_factor5: str="",
                 department: str="", faculty: str=""):

        super().__init__(title, year, authors, doc_type, author, num_citations, document_name, department, faculty)
        self._impact_factor = impact_factor
        self._impact_factor5 = impact_factor5

    @property
    def impact_factor(self):
        return self._impact_factor

    @impact_factor.setter
    def impact_factor(self, value):
        self._impact_factor = value

    @property
    def impact_factor5(self):
        return self._impact_factor5

    @impact_factor5.setter
    def impact_factor5(self, value):
        self._impact_factor5 = value

    @staticmethod
    def read_from_sheet(sheet: worksheet, row: int):
        work_super = Work.read_from_sheet(sheet, row)
        work = WorkWos(author=work_super.author, title=work_super.title, year=work_super.year,
                       authors=work_super.authors, num_citations=work_super.num_citations,
                       doc_type=work_super.document_type, document_name=work_super.document_name,
                       department=work_super.department, faculty=work_super.faculty)
        work.impact_factor = sheet.cell(row, WorkWos.COLUMN_IDX_IMPACT_FACTOR).value
        work.impact_factor5 = sheet.cell(row, WorkWos.COLUMN_IDX_IMPACT_FACTOR5).value
        return work

    def write_headers_to_sheet(self, sheet: worksheet):
        super().write_headers_to_sheet(sheet)
        sheet.cell(1, WorkWos.COLUMN_IDX_IMPACT_FACTOR).value = WorkWos.COLUMN_IMPACT_FACTOR
        sheet.cell(1, WorkWos.COLUMN_IDX_IMPACT_FACTOR5).value = WorkWos.COLUMN_IMPACT_FACTOR5

    def write_to_sheet(self, sheet: worksheet, row: int):
        super().write_to_sheet(sheet, row)
        sheet.cell(row, WorkWos.COLUMN_IDX_IMPACT_FACTOR).value = self.impact_factor
        sheet.cell(row, WorkWos.COLUMN_IDX_IMPACT_FACTOR5).value = self.impact_factor5

