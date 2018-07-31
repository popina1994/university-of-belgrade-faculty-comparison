from openpyxl.worksheet import worksheet


class Author:
    def __init__(self, first_name: str, last_name: str, department: str, faculty: str, middle_name="", link=""):
        self._first_name = first_name
        self._last_name = last_name
        self._department = department
        self._faculty = faculty
        self._middle_name = middle_name
        self._link = link

    COLUMN_FIRST_NAME = "Ime"
    COLUMN_LAST_NAME = "Prezime"
    COLUMN_DEPARTMENT_NAME = "Odsek"
    COLUMN_FACULTY_NAME = "Fakultet"
    COLUMN_MIDDLE_NAME = "Srednje ime"
    COLUMN_LINK_NAME = "Link"

    COLUMN_IDX_FIRST_NAME = 1
    COLUMN_IDX_LAST_NAME = 2
    COLUMN_IDX_MIDDLE_NAME = 3
    COLUMN_IDX_DEPARTMENT_NAME = 4
    COLUMN_IDX_FACULTY_NAME = 5
    COLUMN_IDX_LINK = 6

    MIDDLE_NAME_NOT_FOUND = "N/A"


    @property
    def first_name(self):
        return self._first_name

    @first_name.setter
    def first_name(self, value):
        self._first_name = value

    @property
    def middle_name(self):
        return self._middle_name

    @middle_name.setter
    def middle_name(self, value):
        self._middle_name = value

    @property
    def last_name(self):
        return self._last_name

    @last_name.setter
    def last_name(self, value):
        self._last_name = value

    @property
    def department(self):
        return self._department

    @department.setter
    def department(self, value):
        self._department = value

    @property
    def faculty(self):
        return self._faculty

    @faculty.setter
    def faculty(self, value):
        self._faculty = value

    @property
    def link(self):
        return self._link

    @link.setter
    def link(self, value):
        self._link = value

    def __lt__(self, other):
        if self.last_name != other.last_name:
            return self.last_name < other.last_name
        if self._first_name != other._first_name:
            return self._first_name < other._first_name
        return self.middle_name < other.middle_name

    def __eq__(self, other):
        return (self._first_name == other.first_name) and (self.last_name == other.last_name)\
                and (self.middle_name == other.middle_name)

    def __hash__(self):
        return hash((self.first_name, self.last_name, self.middle_name))

    @staticmethod
    def write_header_to_sheet(sheet: worksheet):
        sheet.cell(1, Author.COLUMN_IDX_FIRST_NAME).value = Author.COLUMN_FIRST_NAME
        sheet.cell(1, Author.COLUMN_IDX_LAST_NAME).value = Author.COLUMN_LAST_NAME
        sheet.cell(1, Author.COLUMN_IDX_MIDDLE_NAME).value = Author.COLUMN_MIDDLE_NAME
        sheet.cell(1, Author.COLUMN_IDX_FACULTY_NAME).value = Author.COLUMN_FACULTY_NAME
        sheet.cell(1, Author.COLUMN_IDX_DEPARTMENT_NAME).value = Author.COLUMN_DEPARTMENT_NAME
        sheet.cell(1, Author.COLUMN_IDX_LINK).value = Author.COLUMN_LINK_NAME

    def write_to_sheet(self, sheet: worksheet, row: int):
        sheet.cell(row, Author.COLUMN_IDX_FIRST_NAME).value = self.first_name
        sheet.cell(row, Author.COLUMN_IDX_LAST_NAME).value = self.last_name
        sheet.cell(row, Author.COLUMN_IDX_MIDDLE_NAME).value = self.middle_name
        sheet.cell(row, Author.COLUMN_IDX_FACULTY_NAME).value = self.faculty
        sheet.cell(row, Author.COLUMN_IDX_DEPARTMENT_NAME).value = self.department
        sheet.cell(row, Author.COLUMN_IDX_LINK).value = self.link

    @staticmethod
    def id_name_static(first_name: str, last_name: str, middle_name: str):
        return "{} {} {}".format(last_name, first_name, middle_name).strip()

    def id_name(self):
        return Author.id_name_static(self.first_name, self.last_name, self.middle_name)

    def __str__(self):
        return (self.first_name + ", " +
                self.last_name + ", " +
                self.department + ", " +
                self.faculty + ", ")
