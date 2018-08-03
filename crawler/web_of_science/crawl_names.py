import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook

from data.workbooks.authors_workbook import AUTHORS_FILE_NAME
from data.author import Author
from utilities.global_setup import DATA_PATH

KOBSON_PATH = "http://kobson.nb.rs/nauka_u_srbiji.133.html?prezime={}+{}%25"
AUTHORS_WOS_FILE_NAME = DATA_PATH + r"\People\authors_wos.xlsx"
AUTHORS_ALL_WOS_FILE_NAME = DATA_PATH + r"\People\authors_all_wos.xlsx"
ALL_AUTHORS_WOS_SHEET_NAME = "Svi"


def crawl_middle_name(first_name: str, last_name: str):
    path = KOBSON_PATH.format(last_name, first_name)
    r = requests.get(path)
    r.encoding = "utf-8"
    data = r.text
    soup = BeautifulSoup(data)
    cnt = 0
    middle_name = ""
    for results in soup.find_all("table", {"class": "type"}):
        for row in results.find_all("td", {"style": "padding-left:11px;"}):
            for full_name_a in row.find('a', href=True):
                full_name = full_name_a.string.strip()
                sub_names = full_name.split(" ")
                if (sub_names[0].lower() != last_name.lower()) or\
                    (sub_names[1].lower() != first_name.lower()):
                    continue
                if sub_names.__len__() == 3:
                    middle_name = sub_names[-1]
                cnt += 1
                print(middle_name)
    print(first_name + " " + last_name + " " + str(cnt))
    if cnt == 0:
        middle_name = "N/A"
    middle_name = "XXX" if cnt > 1 else middle_name
    return middle_name


def update_middle_names():
    work_book = openpyxl.load_workbook(filename=AUTHORS_FILE_NAME)
    for sheet in work_book.worksheets:
        for row in range(2, sheet.max_row + 1):
            first_name = sheet.cell(row, Author.COLUMN_IDX_FIRST_NAME).value
            last_name = sheet.cell(row, Author.COLUMN_IDX_LAST_NAME).value
            middle_name = crawl_middle_name(first_name, last_name.replace(" ", "-"))
            sheet.cell(row, Author.COLUMN_IDX_MIDDLE_NAME).value = middle_name
    work_book.save(AUTHORS_WOS_FILE_NAME)


def get_list_authors():
    work_book_author = openpyxl.load_workbook(filename=AUTHORS_WOS_FILE_NAME)
    list_authors = []
    for sheet in work_book_author.worksheets:
        for row in range(2, sheet.max_row + 1):
            first_name = sheet.cell(row, Author.COLUMN_IDX_FIRST_NAME).value
            last_name = sheet.cell(row, Author.COLUMN_IDX_LAST_NAME).value
            middle_names = sheet.cell(row, Author.COLUMN_IDX_MIDDLE_NAME).value
            middle_names = "" if middle_names is None else middle_names
            department = sheet.cell(row, Author.COLUMN_IDX_DEPARTMENT_NAME).value
            faculty = sheet.cell(row, Author.COLUMN_IDX_FACULTY_NAME).value
            middle_names_sub = middle_names.split(",")
            for middle_name in middle_names_sub:
                author = Author(first_name, last_name, department, faculty, middle_name)
                list_authors.append(author)
    return list_authors


def write_all_authors():
    # Only used for generation of nodes in gephi.
    list_authors = get_list_authors()
    work_book = Workbook()
    work_book.remove(work_book.active)
    all_authors_sheet = work_book.create_sheet(ALL_AUTHORS_WOS_SHEET_NAME)

    Author.write_header_to_sheet(all_authors_sheet)
    row = 2
    for author in list_authors:
        author.write_to_sheet(all_authors_sheet, row)
        row += 1
    work_book.save(AUTHORS_ALL_WOS_FILE_NAME)


if __name__ == "__main__":
    write_all_authors()
