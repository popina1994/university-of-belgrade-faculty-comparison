from utilities.global_setup import PROXY
from bibtexparser.bparser import BibTexParser
import requests
import re
import openpyxl
from crawler.web_of_science.crawl_names import AUTHORS_WOS_FILE_NAME
from data.author import Author
from data.author import Author

REGEX_AUTHOR_STRING = "author\s*=\s*\{[\w ,-.()]+\},"
KOBSON_WOS_BIBTEX = "http://kobson.nb.rs/aspx/wos/export.aspx?autor={}%20{}{}&samoar=&format=BibTeX"


class CrawlerLinksWos:
    idx_gen = 0

    @staticmethod
    def format_middle_name(middle_name: str):
        return "" if middle_name == "" else "%20" + middle_name

    @staticmethod
    def parse_bib_tex(bib_tex: str):
        bib_tex_parser = BibTexParser()
        bib_database = bib_tex_parser.parse(bib_tex)
        return bib_database.entries

    def crawl_works_author(self, first_name: str, last_name: str, middle_name: str):
        path = KOBSON_WOS_BIBTEX.format(last_name, first_name, CrawlerLinksWos.format_middle_name(middle_name))
        print(path)
        r = requests.get(path, proxies=PROXY)
        r.encoding = "utf-8"
        data = r.text
        data_structured = re.sub("(" + REGEX_AUTHOR_STRING + ")", r"{},\r\n\1".format(CrawlerLinksWos.idx_gen), data)
        CrawlerLinksWos.idx_gen += 1
        return CrawlerLinksWos.parse_bib_tex(data_structured)

    def crawl_works(self):
        work_book = openpyxl.load_workbook(filename=AUTHORS_WOS_FILE_NAME)
        for sheet in work_book.worksheets:
            for row in range(2, sheet.max_row):
                first_name = sheet.cell(row, Author.COLUMN_IDX_FIRST_NAME).value
                last_name = sheet.cell(row, Author.COLUMN_IDX_LAST_NAME).value
                middle_names = sheet.cell(row, Author.COLUMN_IDX_MIDDLE_NAME).value
                middle_names = "" if middle_names is None else middle_names
                middle_names_sub = middle_names.split(",")
                cnt = 0
                for middle_name in middle_names_sub:
                    if middle_name != Author.MIDDLE_NAME_NOT_FOUND:
                        works = self.crawl_works_author(first_name, last_name.replace(" ", "-"), middle_name)
                        print(works)
                    else:
                        print("No works available")


if __name__ == "__main__":
    '''
    data = """@ARTICLE{
year={2018},
author={Petrovic-Savic Suzana,Ristic Branko M,Jovanovic Zoran,Matic Aleksandar,Prodanovic Nikola S,Anwer Nabil,Qiao Lihong,Devedzic Goran B},
title={Parametric Model Variability of the Proximal Femoral Sculptural Shape},
journal={INTERNATIONAL JOURNAL OF PRECISION ENGINEERING AND MANUFACTURING},
volume={19},
number={7},
pages={1047-1054},
document_type={Article},
}
    """
    '''
    crawler = CrawlerLinksWos()
    crawler.crawl_works()
    #print(crawl_works_author("jovanovic", "zoran", ""))
    #print(parse_bib_tex(data))
