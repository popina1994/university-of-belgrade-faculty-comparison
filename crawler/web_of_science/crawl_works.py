from bs4 import BeautifulSoup
from fake_useragent import UserAgent

from data.workbooks.graph_edges_workbook import GraphEdgesWorkbook
from data.workbooks.works_workbook import WorksWorkbook, WORKS_WOS_FILE_NAME, WORKS_WOS_SHEET_NAME
from utilities.global_setup import PROXY
from bibtexparser.bparser import BibTexParser
import requests
import re
import openpyxl
from crawler.web_of_science.crawl_names import get_list_authors
from data.author import Author
from data.work import Work
from data.graph_edge import GraphEdge
from math import ceil

REGEX_AUTHOR_STRING = "author\s*=\s*\{[\w ,-.()]+\},"
KOBSON_WOS_BIBTEX = "http://kobson.nb.rs/aspx/wos/export.aspx?autor={}%20{}{}&samoar=&format=BibTeX"
KOBSON_WOS = "http://kobson.nb.rs/nauka_u_srbiji.132.html?autor={}%20{}{}&offset={}"
KOBSON_SERVICE_ROOT = "http://kobson.nb.rs/{}"
WOS_TEXT_CELL = "ISI/Web of Science"
WOS_JOURNAL_RANG_TEXT_CELL = "Rang časopisa"
WOS_JOURNAL_IMPACT_FACTOR_TEXT_CELL = "oblast  / impakt faktor"
NUM_WORKS_PER_PAGE = 10
LAST_YEAR = 2017


class CrawlerLinksWos:
    idx_gen = 0

    def __init__(self):
        self.list_authors = get_list_authors()
        list_name_authors = [author.id_name().lower() for author in self.list_authors]
        self.set_name_authors = set(list_name_authors)
        self.cur_row = 1
        self.user_agent = UserAgent()

    @staticmethod
    def format_middle_name(middle_name: str):
        return "" if middle_name == "" else "%20" + middle_name

    @staticmethod
    def parse_bib_tex(bib_tex: str):
        bib_tex_parser = BibTexParser()
        bib_database = bib_tex_parser.parse(bib_tex)
        return bib_database.entries

    @staticmethod
    def crawl_works_author(first_name: str, last_name: str, middle_name: str):
        path = KOBSON_WOS_BIBTEX.format(last_name, first_name, CrawlerLinksWos.format_middle_name(middle_name))
        r = requests.get(path, proxies=PROXY)
        r.encoding = "utf-8"
        data = r.text
        data_structured = re.sub("(" + REGEX_AUTHOR_STRING + ")", r"{},\r\n\1".format(CrawlerLinksWos.idx_gen), data)
        CrawlerLinksWos.idx_gen += 1
        return path, CrawlerLinksWos.parse_bib_tex(data_structured)

    def get_random_header(self):
        return {"User-Agent": self.user_agent.random}

    def crawl_wos_and_journal_links(self, first_name: str, last_name: str, middle_name:str, number_works: int):
        wos_links = []
        journal_links = []
        num_pages = -(-number_works // NUM_WORKS_PER_PAGE)
        for offset in range(0, num_pages):
            path = KOBSON_WOS.format(last_name, first_name, CrawlerLinksWos.format_middle_name(middle_name), offset)
            links_to_crawl = NUM_WORKS_PER_PAGE if (offset < (num_pages - 1)) \
                                                else number_works - (num_pages - 1) * NUM_WORKS_PER_PAGE
            print(links_to_crawl)
            r = requests.get(path, headers=self.get_random_header())
            r.encoding = "utf-8"
            data = r.text
            soup = BeautifulSoup(data)
            for green_rows in soup.find_all("tr", {"class": "greenRow"}):
                for rCell in green_rows.find_all("td", {"class": "rCell"}):
                    wos_link_added = False
                    journal_link = None
                    for link in rCell.find_all("a", href=True):
                        if (not wos_link_added) and (WOS_TEXT_CELL.lower() == link.text.strip().lower()):
                            wos_links.append(link.get('href'))
                            wos_link_added = True
                        if WOS_JOURNAL_RANG_TEXT_CELL.lower() == link.text.strip().lower():
                            journal_link = KOBSON_SERVICE_ROOT.format(link.get('href'))
                    journal_links.append(journal_link)
                links_to_crawl -= 1
                if links_to_crawl == 0:
                    break

        return journal_links, wos_links

    def get_num_citations(self, wos_link: str):
        r = requests.get(wos_link, headers=self.get_random_header())
        r.encoding = "utf-8"
        data = r.text
        soup = BeautifulSoup(data)
        for citation_div in soup.find_all("div", {"class": "flex-row-partition1"}):
            for num_citation in citation_div.find_all("span", {"class": "large-number"}):
                return num_citation.text

    @staticmethod
    def find_offset_year(table, year: int):
        for row in table.find_all("tr"):
            row_header_elem = row.find("td", {"class": "first dblue"})
            if row_header_elem is not None:
                for offset, cell in enumerate(row.find_all("td", {"class": "dblue"})):
                    if cell.text == str(year):
                        return offset

    @staticmethod
    def get_impact_factor(soup: BeautifulSoup, is_five_year, year: int):
        skip_impact_factor = is_five_year
        skip_header = True
        for data_div in soup.find_all("div", {"class": "resultHolder"}):
            if skip_header:
                skip_header = False
                continue
            for table in data_div.find_all("table", {"class": "type categories results"}):
                if skip_impact_factor:
                    skip_impact_factor = False
                    continue
                offset = CrawlerLinksWos.find_offset_year(table, year)
                if offset is None:
                    return None
                for row in table.find_all("tr"):
                    row_header_elem = row.find("td", {"class": "first lblue"})
                    row_header = "" if row_header_elem is None else row_header_elem.text
                    if row_header.lower() == WOS_JOURNAL_IMPACT_FACTOR_TEXT_CELL.lower():
                        return row.find_all("td", {"class": "lblue"})[offset].text

    def get_impact_factors(self, journal_link: str, year: int):
        if journal_link is None:
            return "", ""
        r = requests.get(journal_link, headers=self.get_random_header())
        r.encoding = "utf-8"
        data = r.text
        soup = BeautifulSoup(data)
        impact_factor_2017 = CrawlerLinksWos.get_impact_factor(soup, False, year)
        impact_factor_2017 = "" if impact_factor_2017 is None else impact_factor_2017
        impact_factor5_2017 = CrawlerLinksWos.get_impact_factor(soup, True, year)
        impact_factor5_2017 = "" if impact_factor5_2017 is None else impact_factor5_2017
        return impact_factor_2017, impact_factor5_2017

    def crawl_works(self):
        works_work_book = WorksWorkbook()

        for author in self.list_authors:
            print("{} {} {}".format(author.first_name, author.last_name, author.middle_name))
            if author.middle_name != Author.MIDDLE_NAME_NOT_FOUND:
                path, works = CrawlerLinksWos.crawl_works_author(author.first_name,
                                                                 author.last_name.replace(" ", "-"),
                                                                 author.middle_name)
                journal_links, wos_links = self.crawl_wos_and_journal_links(
                                            author.first_name, author.last_name.replace(" ", "-"),
                                            author.middle_name, works.__len__())
                print("Number of works {}".format(works.__len__()))
                for work_id, work_bib in enumerate(works):
                    impact_factor, impact_factor5 = self.get_impact_factors(journal_links[work_id], LAST_YEAR)
                    print(impact_factor5)
                    num_citations = self.get_num_citations(wos_links[work_id])
                    print("if{} if5{} num_cit {}".format(impact_factor, impact_factor5, num_citations))
                    work = Work(title=work_bib["title"], authors=work_bib["author"].replace("-", " "),
                                year=work_bib["year"], doc_type=work_bib['document_type'],
                                author="{} {} {}".format(author.last_name, author.first_name, author.middle_name).strip(),
                                num_citations=num_citations,
                                document_name=work_bib['journal'],
                                impact_factor=impact_factor, impact_factor5=impact_factor5,
                                department=author.department, faculty=author.faculty)
                    works_work_book.save_work(work)
                    print(work_bib["title"])
            else:
                print("No works available")
        works_work_book.save()

    def generate_graph_known_authors(self):
        work_book_works = openpyxl.load_workbook(filename=WORKS_WOS_FILE_NAME)
        sheet = work_book_works[WORKS_WOS_SHEET_NAME]
        work_book_edges = GraphEdgesWorkbook()
        for row in range(2, sheet.max_row + 1):
            author1 = sheet.cell(row, Work.COLUMN_IDX_AUTHOR).value.lower()
            authors = sheet.cell(row, Work.COLUMN_IDX_AUTHORS).value.lower()
            for author2 in authors.split(","):
                if author2 in self.set_name_authors:
                    edge = GraphEdge(author1, author2)
                    work_book_edges.save_graph_edge(edge)
        work_book_edges.save()

    def generate_graph_authors(self):
        pass


if __name__ == "__main__":
    crawler = CrawlerLinksWos()
    crawler.crawl_works()
    #crawler.generate_graph_known_authors()
    #print(crawl_works_author("jovanovic", "zoran", ""))
    #print(parse_bib_tex(data))
