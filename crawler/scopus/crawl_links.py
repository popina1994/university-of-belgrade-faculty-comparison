import requests
from bs4 import BeautifulSoup
import openpyxl
from crawler.web_of_science.crawl_names import  AUTHORS_WOS_FILE_NAME
from data.tables.author import Author
from utilities.global_setup import PROXY

SCOPUS_PATH = "https://www.scopus.com/results/authorNamesList.uri?sort=count-f&src=al&sid=8a5f928c9f2a3530c144238b9d2f5e02&sot=al&sdt=al&sl=74&s=AUTHLASTNAME%28{}%29+AND+AUTHFIRST%28{}%29&st1={}&st2={}&orcidId=&selectionPageSearch=anl&reselectAuthor=false&activeFlag=true&showDocument=false&resultsPerPage=20&offset=1&jtp=false&currentPage=1&previousSelectionCount=0&tooManySelections=false&previousResultCount=0&authSubject=LFSC&authSubject=HLSC&authSubject=PHSC&authSubject=SOSC&exactAuthorSearch=false&showFullList=false&authorPreferredName=&origin=searchauthorfreelookup&affiliationId=&txGid=855cc953ff2e66c64f5d95ddb748ce36"
AUTHORS_SCOPUS_FILE_NAME = r"C:\Users\popina\Dropbox\Fakultet\Master Thesis\Data\People\authors_scopus.xlsx"


def crawl_links(first_name: str, last_name: str):
    path = SCOPUS_PATH.format(last_name, first_name, last_name, first_name)
    r = requests.get(path, proxies=PROXY)
    r.encoding = "utf-8"
    data = r.text
    soup = BeautifulSoup(data)
    links = []
    print(first_name + " " + last_name)
    for results in soup.find_all("td", {"class": "authorResultsNamesCol col20"}):
        for full_name_a in results.find_all('a', href=True):
            link = full_name_a['href']
            links.append(link)
            print(link)
    return links


def update_links():
    work_book = openpyxl.load_workbook(filename=AUTHORS_WOS_FILE_NAME)
    for sheet in work_book.worksheets:
        sheet.cell(1, Author.COLUMN_IDX_LINK).value = Author.COLUMN_LINK_NAME
        for row in range(2, sheet.max_row + 1):
            first_name = sheet.cell(row, Author.COLUMN_IDX_FIRST_NAME).value
            last_name = sheet.cell(row, Author.COLUMN_IDX_LAST_NAME).value
            crawled_links = crawl_links(first_name, last_name.replace(" ", "-"))
            link_str = ""
            first = True
            for it in crawled_links:
                if first:
                    link_str = it
                    first = False
                else:
                    link_str += "\n, " + it
            sheet.cell(row, Author.COLUMN_IDX_LINK).value = link_str
    #work_book.save(AUTHORS_SCOPUS_FILE_NAME)


def get_list_authors():
    work_book_author = openpyxl.load_workbook(filename=AUTHORS_SCOPUS_FILE_NAME)
    list_authors = []
    for sheet in work_book_author.worksheets:
        for row in range(2, sheet.max_row + 1):
            first_name = sheet.cell(row, Author.COLUMN_IDX_FIRST_NAME).value
            last_name = sheet.cell(row, Author.COLUMN_IDX_LAST_NAME).value
            middle_names = sheet.cell(row, Author.COLUMN_IDX_MIDDLE_NAME).value
            middle_names = "" if middle_names is None else middle_names
            department = sheet.cell(row, Author.COLUMN_IDX_DEPARTMENT_NAME).value
            faculty = sheet.cell(row, Author.COLUMN_IDX_FACULTY_NAME).value
            link_v = sheet.cell(row, Author.COLUMN_IDX_LINK).value
            link = "" if link_v is None else link_v
            middle_name = middle_names.split(",")[0]
            author = Author(first_name, last_name, department, faculty, middle_name, link)
            list_authors.append(author)
    return list_authors


if __name__ == "__main__":
    print(get_list_authors())
