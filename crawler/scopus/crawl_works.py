import os
import shutil
import time

from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

from data.workbooks.graph_edges_workbook import GraphEdgesWorkbook
from data.workbooks.works_workbook import WorksWorkbook, WORKS_WOS_FILE_NAME, WORKS_SHEET_NAME
from utilities.global_setup import PROXY, SELENIUM_CHROME_DRIVER_PATH, DATA_PATH
from bibtexparser.bparser import BibTexParser
import requests
import openpyxl
from crawler.scopus.crawl_links import get_list_authors
from data.author import Author
from data.work import Work
from data.graph_edge import GraphEdge
from selenium import webdriver
from bibtexparser.customization import convert_to_unicode

REGEX_AUTHOR_STRING = "author\s*=\s*\{[\w ,-.()]+\},"
DOWNLOAD_FILE_PATH = DATA_PATH + "\\Work\\scopus"
SCOPUS_ROOT = 'https://www.scopus.com'
TIME_LIMIT_WAIT = 60


class CrawlerLinksScopus:
    idx_gen = 0

    def __init__(self):
        self.list_authors = get_list_authors()
        list_name_authors = [author.id_name().lower() for author in self.list_authors]
        self.set_name_authors = set(list_name_authors)
        self.cur_row = 1
        self.user_agent = UserAgent()
        self.driver = webdriver.Chrome(SELENIUM_CHROME_DRIVER_PATH)

    @staticmethod
    def format_middle_name(middle_name: str):
        return "" if middle_name == "" else "%20" + middle_name

    @staticmethod
    def parse_bib_tex_file(path: str):
        bib_tex_parser = BibTexParser()
        bib_tex_parser.customization = convert_to_unicode
        with open(path, encoding="utf-8") as bib_tex_file:
            bib_database = bib_tex_parser.parse_file(bib_tex_file)
        return bib_database.entries

    @staticmethod
    def init_driver_for_works_by_author(download_dir):
        if os.path.exists(download_dir):
            shutil.rmtree(download_dir)
        os.makedirs(download_dir)
        options = webdriver.ChromeOptions()
        prefs = {"download.default_directory": download_dir}
        options.add_experimental_option("prefs", prefs)
        driver = webdriver.Chrome(SELENIUM_CHROME_DRIVER_PATH, chrome_options=options)
        return driver

    @staticmethod
    def wait_and_click(driver, x_path):
        webdriver_wait = WebDriverWait(driver, TIME_LIMIT_WAIT).until(
            expected_conditions.presence_of_all_elements_located((By.XPATH, x_path)))
        time.sleep(1)
        element = driver.find_element_by_xpath(x_path)
        element.click()

    @staticmethod
    def download_bibtex(driver, first_download: bool):
        SELECT_BOX_X_PATH = '//*[@id="export_results"]/span'
        STUPID_POP_UP_X_PATH = '//*[@id="_pendo-close-guide_"]'
        RADIO_BUTTON_X_PATH = '//*[@id="exportList"]/li[5]'
        EXPORT_BUTTON_X_PATH = '//*[@id="exportTrigger"]'
        CrawlerLinksScopus.wait_and_click(driver, SELECT_BOX_X_PATH)
        try:
            if first_download:
                CrawlerLinksScopus.wait_and_click(driver, STUPID_POP_UP_X_PATH)
        except:
            pass
        CrawlerLinksScopus.wait_and_click(driver, RADIO_BUTTON_X_PATH)
        CrawlerLinksScopus.wait_and_click(driver, EXPORT_BUTTON_X_PATH)

    @staticmethod
    def get_num_citation_for_document(row):
        td = row.find_all("td", {"class": "textRight"})[-1]
        ahref = td.find("a")
        cite = 0 if ahref is None else ahref.text.strip()
        return cite

    @staticmethod
    def get_field_weight_link_for_document(row):
        td = row.find_all("td")[0]
        ahref = td.find("a")
        field_weight_link = "" if ahref is None else ahref.get('href')
        return field_weight_link

    @staticmethod
    def get_journal_link_for_document(row):
        td = row.find_all("td")[-2]
        ahref = td.find("a")
        journal_link = "" if ahref is None else SCOPUS_ROOT + ahref.get('href')
        return journal_link

    @staticmethod
    def get_num_citations_and_field_weight_links_and_journal_links(driver):
        SELECT_MENU_X_PATH = '//*[@id="resultsPerPage-button"]/span[1]'
        SELECT_OPTION_X_PATH = '//*[@id="resultsPerPage-menu"]/li[4]'
        CrawlerLinksScopus.wait_and_click(driver, SELECT_MENU_X_PATH)
        CrawlerLinksScopus.wait_and_click(driver, SELECT_OPTION_X_PATH)
        time.sleep(5)
        data = driver.page_source
        soup = BeautifulSoup(data)

        citations = []
        field_weight_links = []
        journal_links = []
        for row in soup.find_all("tr", {"class": "searchArea"}):
            cite = CrawlerLinksScopus.get_num_citation_for_document(row)
            citations.append(cite)
            field_weight_link = CrawlerLinksScopus.get_field_weight_link_for_document(row)
            field_weight_links.append(field_weight_link)
            journal_link = CrawlerLinksScopus.get_journal_link_for_document(row)
            journal_links.append(journal_link)
        return citations, field_weight_links, journal_links

    @staticmethod
    def crawl_works_citations_and_field_weight_links_author(author: Author):
        download_dir = "{}\\{}_{}".format(DOWNLOAD_FILE_PATH, author.last_name, author.first_name)
        driver = CrawlerLinksScopus.init_driver_for_works_by_author(download_dir)
        works = []
        citations_total = []
        field_weight_links_total = []
        journal_links_total = []
        print(author.first_name + " " + author.last_name)
        first_download = True

        for path in author.link.split(","):
            path = path.strip()
            print(path)
            driver.get(path)
            CrawlerLinksScopus.download_bibtex(driver, first_download)
            first_download = False

            citations, field_weight_links, journal_links = \
                CrawlerLinksScopus.get_num_citations_and_field_weight_links_and_journal_links(driver)
            citations_total += citations
            field_weight_links_total += field_weight_links
            journal_links_total += journal_links

        # Latency of download
        time.sleep(.3)
        for file in os.listdir(download_dir):
            works += CrawlerLinksScopus.parse_bib_tex_file(os.path.join(download_dir, file))
        return works, citations_total, field_weight_links_total, journal_links_total

    @staticmethod
    def get_weight_index(document_link: str):
        if document_link is "":
            return ""
        success = False
        field_weighted_citation_impact = ""
        while not success:
            driver = webdriver.Chrome(SELENIUM_CHROME_DRIVER_PATH)
            driver.get(document_link)
            time.sleep(2)
            soup = BeautifulSoup(driver.page_source)
            try:
                field_weighted_citation_impact = soup.find_all("div", {"class": "metricCount"})[1].text
                success = True
            except:
                success = False

        return field_weighted_citation_impact

    def get_journal_factors(self, journal_link: str):
        if journal_link is "":
            return ""
        print(journal_link)
        success = False
        while not success:
            #driver = webdriver.Chrome(SELENIUM_CHROME_DRIVER_PATH)
            self.driver.get(journal_link)
            time.sleep(2)
            soup = BeautifulSoup(self.driver.page_source)
            try:
                cite_score = soup.find_all("div", {"class": ["value", "fontMedLarge", "lineHeight2"]})[0].text
                sjr = soup.find_all("div", {"class": ["value", "fontMedLarge", "lineHeight2"]})[1].text
                snip = soup.find_all("div", {"class": ["value", "fontMedLarge", "lineHeight2"]})[2].text
                success = True
            except:
                success = False

        return cite_score, sjr, snip

    def crawl_works(self):
        works_work_book = WorksWorkbook(is_wos=False)

        for author in self.list_authors:
            if author.link != "":
                works, citations, field_weight_links, journal_links  = \
                    CrawlerLinksScopus.crawl_works_citations_and_field_weight_links_author(author)
                print("Number of works {}".format(works.__len__()))
                for work_id, work_bib in enumerate(works):

                    #impact_factor = CrawlerLinksScopus.get_weight_index(field_weight_links[work_id])
                    cite_factor, sjr, snip = self.get_journal_factors(journal_links[work_id])

                    print("cf{} sjr{} snip{} num_cit {}".format(cite_factor, sjr, snip, citations[work_id]))

                    work = Work(title=work_bib["title"], authors=work_bib["author"].replace("-", " "),
                                year=work_bib["year"], doc_type=work_bib['document_type'],
                                author="{} {} {}".format(author.last_name, author.first_name, author.middle_name).strip(),
                                num_citations=snip,
                                document_name=work_bib.get('journal', ""),
                                impact_factor=cite_factor, impact_factor5=sjr,
                                department=author.department, faculty=author.faculty)
                    works_work_book.save_work(work)
                    print(work_bib["title"])
            else:
                print("No works available")
        works_work_book.save()

    def generate_graph_known_authors(self):
        work_book_works = openpyxl.load_workbook(filename=WORKS_WOS_FILE_NAME)
        sheet = work_book_works[WORKS_SHEET_NAME]
        work_book_edges = GraphEdgesWorkbook()
        for row in range(2, sheet.max_row + 1):
            author1 = sheet.cell(row, Work.COLUMN_IDX_AUTHOR).value.lower()
            authors = sheet.cell(row, Work.COLUMN_IDX_AUTHORS).value.lower()
            for author2 in authors.split(","):
                if author2 in self.set_name_authors:
                    edge = GraphEdge(author1, author2)
                    work_book_edges.save_graph_edge(edge)
        work_book_edges.save()

    def generate_graph_authors(self):
        pass


if __name__ == "__main__":
    crawler = CrawlerLinksScopus()
    crawler.crawl_works()
