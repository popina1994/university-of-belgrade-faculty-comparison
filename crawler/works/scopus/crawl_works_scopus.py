import os
import shutil
import time

from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

from crawler.authors.crawler_matf import MATF_DEPARTMENT, MATF_FACULTY_NAME
from crawler.works.crawl_works import CrawlerWorks
from crawler.works.scopus.crawl_links import get_list_authors
from data.tables.work_table.work_scopus import WorkScopus
from data.workbooks.works_workbook import WorksWorkbook, WORKS_SHEET_NAME, WORKS_FILE_NAME, \
    WorkTypes
from language.language_converter import CyrillicLatin
from utilities.global_setup import SELENIUM_CHROME_DRIVER_PATH, DATA_PATH
from bibtexparser.bparser import BibTexParser
import openpyxl
from data.tables.author import Author
from data.tables.work_table.work import Work
from selenium import webdriver
from bibtexparser.customization import convert_to_unicode

REGEX_AUTHOR_STRING = "author\s*=\s*\{[\w ,-.()]+\},"
DOWNLOAD_FILE_PATH = DATA_PATH + "\\Work\\scopus"
SCOPUS_ROOT = 'https://www.scopus.com'
TIME_LIMIT_WAIT = 60


class CrawlerWorksScopus(CrawlerWorks):
    def __init__(self):
        super().__init__(WorkTypes.SCOPUS)
        self.driver = webdriver.Chrome(SELENIUM_CHROME_DRIVER_PATH)

    def get_list_authors(self):
        return get_list_authors()

    @staticmethod
    def parse_bib_tex_file(path: str):
        bib_tex_parser = BibTexParser()
        bib_tex_parser.customization = convert_to_unicode
        with open(path, encoding="utf-8") as bib_tex_file:
            bib_database = bib_tex_parser.parse_file(bib_tex_file)
        return bib_database.entries

    @staticmethod
    def init_driver_for_works_by_author(download_dir):
        if os.path.exists(download_dir):
            shutil.rmtree(download_dir)
        os.makedirs(download_dir)
        options = webdriver.ChromeOptions()
        prefs = {"download.default_directory": download_dir}
        options.add_experimental_option("prefs", prefs)
        driver = webdriver.Chrome(SELENIUM_CHROME_DRIVER_PATH, chrome_options=options)
        return driver

    @staticmethod
    def wait_and_click(driver, x_path):
        webdriver_wait = WebDriverWait(driver, TIME_LIMIT_WAIT).until(
            expected_conditions.presence_of_all_elements_located((By.XPATH, x_path)))
        time.sleep(1)
        element = driver.find_element_by_xpath(x_path)
        element.click()

    @staticmethod
    def download_bibtex(driver, first_download: bool):
        SELECT_BOX_X_PATH = '//*[@id="export_results"]/span'
        STUPID_POP_UP_X_PATH = '//*[@id="_pendo-close-guide_"]'
        RADIO_BUTTON_X_PATH = '//*[@id="exportList"]/li[5]'
        EXPORT_BUTTON_X_PATH = '//*[@id="exportTrigger"]'
        CrawlerWorksScopus.wait_and_click(driver, SELECT_BOX_X_PATH)
        try:
            if first_download:
                CrawlerWorksScopus.wait_and_click(driver, STUPID_POP_UP_X_PATH)
        except:
            pass
        CrawlerWorksScopus.wait_and_click(driver, RADIO_BUTTON_X_PATH)
        CrawlerWorksScopus.wait_and_click(driver, EXPORT_BUTTON_X_PATH)

    @staticmethod
    def get_num_citation_for_document(row):
        td = row.find_all("td", {"class": "textRight"})[-1]
        ahref = td.find("a")
        cite = 0 if ahref is None else ahref.text.strip()
        return cite

    @staticmethod
    def get_field_weight_link_for_document(row):
        td = row.find_all("td")[0]
        ahref = td.find("a")
        field_weight_link = "" if ahref is None else ahref.get('href')
        return field_weight_link

    @staticmethod
    def get_journal_link_for_document(row):
        td = row.find_all("td")[-2]
        ahref = td.find("a")
        journal_link = "" if ahref is None else SCOPUS_ROOT + ahref.get('href')
        return journal_link

    @staticmethod
    def get_num_citations_and_field_weight_links_and_journal_links(driver):
        SELECT_MENU_X_PATH = '//*[@id="resultsPerPage-button"]/span[1]'
        SELECT_OPTION_X_PATH = '//*[@id="resultsPerPage-menu"]/li[4]'
        CrawlerWorksScopus.wait_and_click(driver, SELECT_MENU_X_PATH)
        CrawlerWorksScopus.wait_and_click(driver, SELECT_OPTION_X_PATH)
        time.sleep(5)
        data = driver.page_source
        soup = BeautifulSoup(data)

        citations = []
        field_weight_links = []
        journal_links = []
        for row in soup.find_all("tr", {"class": "searchArea"}):
            cite = CrawlerWorksScopus.get_num_citation_for_document(row)
            citations.append(cite)
            field_weight_link = CrawlerWorksScopus.get_field_weight_link_for_document(row)
            field_weight_links.append(field_weight_link)
            journal_link = CrawlerWorksScopus.get_journal_link_for_document(row)
            journal_links.append(journal_link)
        return citations, field_weight_links, journal_links

    @staticmethod
    def crawl_works_citations_and_field_weight_links_author(author: Author):
        download_dir = "{}\\{}_{}".format(DOWNLOAD_FILE_PATH, author.last_name, author.first_name)
        driver = CrawlerWorksScopus.init_driver_for_works_by_author(download_dir)
        works = []
        citations_total = []
        field_weight_links_total = []
        journal_links_total = []
        print(author.first_name + " " + author.last_name)
        first_download = True

        for path in author.link.split(","):
            path = path.strip()
            print(path)
            driver.get(path)
            CrawlerWorksScopus.download_bibtex(driver, first_download)
            first_download = False

            citations, field_weight_links, journal_links = \
                CrawlerWorksScopus.get_num_citations_and_field_weight_links_and_journal_links(driver)
            citations_total += citations
            field_weight_links_total += field_weight_links
            journal_links_total += journal_links

        # Latency of download
        time.sleep(.3)
        for file in os.listdir(download_dir):
            works += CrawlerWorksScopus.parse_bib_tex_file(os.path.join(download_dir, file))
        return works, citations_total, field_weight_links_total, journal_links_total

    def get_weight_index(self, document_link: str):
        if document_link is "":
            return ""
        success = False
        field_weighted_citation_impact = ""
        while not success:
            self.driver.get(document_link)
            time.sleep(2)
            soup = BeautifulSoup(self.driver.page_source)
            try:
                field_weighted_citation_impact = soup.find_all("div", {"class": "metricCount"})[1].text
                success = True
            except:
                success = False

        return field_weighted_citation_impact

    def get_journal_factors(self, journal_link: str):
        if journal_link is "":
            return "", "", ""
        print(journal_link)
        success = False
        cite_score = ""
        sjr = ""
        snip = ""
        while not success:
            self.driver.get(journal_link)
            time.sleep(2)
            soup = BeautifulSoup(self.driver.page_source)
            try:
                cite_score = soup.find_all("div", {"class": ["value", "fontMedLarge", "lineHeight2"]})[0].text
                sjr = soup.find_all("div", {"class": ["value", "fontMedLarge", "lineHeight2"]})[1].text
                snip = soup.find_all("div", {"class": ["value", "fontMedLarge", "lineHeight2"]})[2].text
                success = True
            except:
                break

        return cite_score, sjr, snip

    def crawl_works(self, list_authors: list, work_book_type: WorkTypes):
        works_work_book = WorksWorkbook(work_book_type)

        for author in list_authors:
            if author.link != "":
                works, citations, field_weight_links, journal_links = \
                    CrawlerWorksScopus.crawl_works_citations_and_field_weight_links_author(author)
                print("Number of works {}".format(works.__len__()))
                for work_id, work_bib in enumerate(works):
                    weight_index = self.get_weight_index(field_weight_links[work_id])
                    cite_score, sjr, snip = self.get_journal_factors(journal_links[work_id])
                    print("cf{} sjr{} snip{} num_cit {}, weight_index {}".
                          format(cite_score, sjr, snip, citations[work_id], weight_index))

                    work = WorkScopus(title=work_bib["title"], authors=work_bib["author"].replace("-", " "),
                                      year=work_bib["year"], doc_type=work_bib['document_type'],
                                      author="{} {} {}".format(author.last_name, author.first_name,
                                      author.middle_name).strip(),
                                      num_citations=citations[work_id], weight_index=weight_index,
                                      document_name=work_bib.get('journal', ""),
                                      cite_score=cite_score, sjr=sjr, snip=snip,
                                      department=author.department, faculty=author.faculty)
                    works_work_book.save_work(work)
                    print(work_bib["title"])
            else:
                print("No works available")
        works_work_book.save()
        crawler.convert_authors_to_real_names(work_book_type)

    def find_author(self, last_name: str, first_name_initial: str):
        for author_it in self.list_authors:
            if (author_it.last_name.lower() == last_name) and \
               (first_name_initial.lower() == author_it.first_name[0].lower()) and (author_it.link != ""):
                return author_it

    def convert_authors_to_real_names(self, work_book_type: WorkTypes):
        work_book_name = WORKS_FILE_NAME[work_book_type]
        work_book_works = openpyxl.load_workbook(filename=work_book_name)
        sheet = work_book_works[WORKS_SHEET_NAME]
        for row in range(2, sheet.max_row + 1):
            authors = sheet.cell(row, Work.COLUMN_IDX_AUTHORS).value.lower()
            authors_new = ""
            first = True
            for author in authors.split(" and "):
                author_sub_names = author.split(",")
                try:
                    last_name = CyrillicLatin.convert_serb_latin_to_latin(author_sub_names[0].strip())
                    first_name_initial_exist = author_sub_names.__len__() > 1
                    first_name_initial = "" if not first_name_initial_exist  \
                                        else CyrillicLatin.convert_serb_latin_to_latin(author_sub_names[1].strip()[0])
                    author_search = self.find_author(last_name, first_name_initial)
                    author_new = author_search.id_name() if author_search is not None \
                        else last_name + " " + first_name_initial + " DOT"
                    if not first:
                        authors_new += ","
                    first = False
                except:
                    print("Test")
                authors_new += author_new

            sheet.cell(row, Work.COLUMN_IDX_AUTHORS).value = authors_new
        work_book_works.save(work_book_name)

    def generate_graph_authors(self):
        pass


if __name__ == "__main__":
    crawler = CrawlerWorksScopus()
    '''
    crawler.crawl_custom_authors([Author(first_name="Sana", last_name="Stojanovic", department=MATF_DEPARTMENT,
                                       faculty=MATF_FACULTY_NAME, middle_name="N",
                                       link=r"https://www.scopus.com/authid/detail.uri?authorId=54401813300")])
'''
    crawler.generate_graph_known_authors_custom()
