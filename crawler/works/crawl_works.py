from abc import ABC, abstractmethod

import openpyxl

from data.tables.graph_edge import GraphEdge
from data.tables.work_table.work import Work
from data.workbooks.authors_all_workbook import AuthorsAllWorkBook
from data.workbooks.graph_edges_workbook import GraphEdgesWorkbook
from data.workbooks.works_workbook import WORKS_FILE_NAME, WORKS_SHEET_NAME, WorkTypes


class CrawlerWorks(ABC):
    def __init__(self, work_book_type: WorkTypes):
        self.work_book_type = work_book_type
        self.list_authors = self.get_list_authors()
        list_name_authors = [author.id_name().lower() for author in self.list_authors]
        self.set_name_authors = set(list_name_authors)

    @abstractmethod
    def get_list_authors(self):
        pass

    def write_all_authors(self):
        work_book = AuthorsAllWorkBook(self.work_book_type)
        for author in self.list_authors:
            work_book.save_author(author)
        work_book.save()

    @abstractmethod
    def crawl_works(self, list_authors: list, work_book_type: WorkTypes):
        pass

    def crawl_work_all_authors(self):
        self.crawl_works(self.list_authors, self.work_book_type)

    def crawl_custom_authors(self, authors: list):
        self.crawl_works(authors, WorkTypes.TEMPORARY)

    def generate_graph_known_authors(self, work_book_type: WorkTypes):
        work_book_works = openpyxl.load_workbook(filename=WORKS_FILE_NAME[work_book_type])
        sheet = work_book_works[WORKS_SHEET_NAME]
        work_book_edges = GraphEdgesWorkbook(work_book_type)
        for row in range(2, sheet.max_row + 1):
            author1 = sheet.cell(row, Work.COLUMN_IDX_AUTHOR).value.lower()
            authors = sheet.cell(row, Work.COLUMN_IDX_AUTHORS).value.lower()
            for author2 in authors.split(","):
                if (author1 < author2) and (author2 in self.set_name_authors):
                    edge = GraphEdge(author1, author2, 1)
                    work_book_edges.save_graph_edge(edge)
        work_book_edges.save()

    def generate_graph_all_known_authors(self):
        self.generate_graph_known_authors(self.work_book_type)

    def generate_graph_known_authors_custom(self):
        self.generate_graph_known_authors(WorkTypes.TEMPORARY)



if __name__ == "__main__":
    pass
    #crawler = CrawlerLinksWos()
    #crawler.crawl_works()
    #crawler.generate_graph_known_authors()